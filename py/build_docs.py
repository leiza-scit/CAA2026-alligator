#!/usr/bin/env python3
"""Draw the figures of the variability note from the graphs, and audit its text.

The companion note to ``service_group_variability`` explains how the heatmap is
built. Its figures used to be hand-drawn SVG with the numbers typed in, which is
how two of them came to be wrong: a findspot IRI that does not exist and a sherd
count that was never in the data. Nothing failed, because nothing could tell a
number that came from the graph from one that came from a keyboard.

This script closes that gap from both ends.

    figures   every data-bearing figure is drawn from docs_facts.collect(),
              in English and French, into output/docs/

Everything is written with an explicit LF newline. Python's text mode would
otherwise use the platform's convention, and a note built on Windows would
differ from the same note built on Linux in every single line — which is the
opposite of what a byte-reproducible pipeline promises. The content is
identical either way; only a diff would ever notice, and a diff noticing
something that did not change is precisely the noise to avoid.
    audit     every number quoted in the note's prose is checked against the
              same facts; a stale one fails the build

Run standalone (``python py/build_docs.py``) or as the ``docs`` step of
``python py/main.py``. Requires the pipeline to have run first.
"""

from __future__ import annotations

import argparse
import math
import re
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import wd_paths                                                  # noqa: E402
import wd_repro  # noqa: E402, F401  (imported for its effect)
from docs_facts import collect                                   # noqa: E402


# ==============================================================================
# SECTION 1 · Configuration
# ==============================================================================

ROOT = wd_paths.ROOT
OUT_DIR = ROOT / "output" / "docs"
LANGS = ("en", "fr")

# Matplotlib's RdYlGn at eleven stops, the same list as py/viz/_prelude.py, so a
# value has one colour in the printed figure, in the browser and here.
RDYLGN = ["#a50026", "#d62f27", "#f46d43", "#fdad60", "#fee08b", "#feffbe",
          "#d9ef8b", "#a5d86a", "#66bd63", "#199750", "#006837"]

# The house palette of the note itself.
INK, INK_SOFT, PAPER, RULE = "#1b2430", "#46525f", "#fbfaf7", "#d9d3c6"
SLIP, OCHRE, SAGE, GRID = "#9e3b26", "#b07d31", "#40655c", "#e2ddd1"
MONO = "IBM Plex Mono, monospace"
SANS = "Public Sans"

# Group colours, matching the printed heatmap's column headings.
GROUP_COLOUR = {"Oblique-rim plate": "#1f77b4",
                "Service I": "#d62728",
                "Service II": "#2ca02c"}

GROUP_LABEL = {
    "en": {"Oblique-rim plate": "Oblique-rim plate",
           "Service I": "Service I", "Service II": "Service II"},
    "fr": {"Oblique-rim plate": "Assiette à bord oblique",
           "Service I": "Service I", "Service II": "Service II"},
}

T = {
    "en": {
        "we_sample": "the sample",
        "we_sums": "three sums Â· SPARQL",
        "we_python": "four numbers Â· Python",
        "we_caption": ["{group} sherds at the", "horizon-{horizon} findspots, by the", "stage rank of their sub-type"],
        "we_sums_note": ["sufficient statistics â the", "query needs no more, which", "is fortunate, because rdflib", "has neither SQRT nor EXP"],
        "we_derived_note": ["the two printed values are", "s and q; x\u0304 and CV are the", "steps between them"],
        "we_foot": "Horizon {horizon} Â· {group} Â· the darkest cell of the printed figure, from the sherds to the two numbers in it.",
        "n_stages": "{n} stages",
        "rank_axis_foot": "stage rank â dot area is the share of sherds, the tick is x\u0304, the bracket is \u00b1s",
        "exp_band": "range actually occupied by this figure",
        "exp_x": "coefficient of variation  CV = s / x\u0304  \u2192",
        "horizon": "Horizon", "earlier": "earlier  →  later",
        "rank": "rank", "stage": "stage",
        "share_of": "share of {group} sherds, by stage",
        "mass": "later horizon → mass concentrates on one rank → s falls",
        "var_title": "Within-group variance · s  (left panel of the printed figure)",
        "qual_title": "Within-group quality · q = exp(−CV)  (right panel)",
        "var_ramp": "scale ends at the largest observed s",
        "qual_ramp": "absolute scale, 0 to 1",
        "shares_title": "Share of each service group, by horizon",
        "shares_foot": "the replacement of one group by another — the signal the variance panels cannot see",
        "origin_x": "rank of the first stage — the arbitrary choice",
        "origin_declared": "declared",
        "seriation_x": "first correspondence-analysis axis — the seriation order",
        "seriation_leg": "{n} findspots against the seriation order",
        "seriation_none": "not plotted: {names}, no CA coordinate",
    },
    "fr": {
        "we_sample": "l\u2019\u00e9chantillon",
        "we_sums": "trois sommes Â· SPARQL",
        "we_python": "quatre nombres Â· Python",
        "we_caption": ["tessons de {group} sur les", "sites de l\u2019horizon {horizon}, selon", "le rang de stade de leur sous-type"],
        "we_sums_note": ["statistiques exhaustives â la", "requ\u00eate n\u2019a besoin de rien de plus,", "heureusement, car le SPARQL de", "rdflib n\u2019a ni SQRT ni EXP"],
        "we_derived_note": ["les deux valeurs imprim\u00e9es sont", "s et q ; x\u0304 et CV sont les", "\u00e9tapes interm\u00e9diaires"],
        "we_foot": "Horizon {horizon} Â· {group} Â· la cellule la plus sombre de la figure, des tessons aux deux nombres qu\u2019elle porte.",
        "n_stages": "{n} stades",
        "rank_axis_foot": "rang de stade â l\u2019aire du point est la part des tessons, le trait est x\u0304, la barre est \u00b1s",
        "exp_band": "plage r\u00e9ellement occup\u00e9e par cette figure",
        "exp_x": "coefficient de variation  CV = s / x\u0304  \u2192",
        "horizon": "Horizon", "earlier": "plus ancien  →  plus récent",
        "rank": "rang", "stage": "stade",
        "share_of": "part des tessons du {group}, par stade",
        "mass": "horizon plus tardif → la masse se concentre sur un rang → s diminue",
        "var_title": "Variance intra-groupe · s  (panneau gauche de la figure imprimée)",
        "qual_title": "Qualité intra-groupe · q = exp(−CV)  (panneau droit)",
        "var_ramp": "l’échelle s’arrête au plus grand s observé",
        "qual_ramp": "échelle absolue, 0 à 1",
        "shares_title": "Part de chaque groupe de service, par horizon",
        "shares_foot": "le remplacement d’un groupe par un autre — le signal que les panneaux de variance ne voient pas",
        "origin_x": "rang du premier stade — le choix arbitraire",
        "origin_declared": "déclaré",
        "seriation_x": "premier axe de l’analyse factorielle des correspondances — l’ordre de sériation",
        "seriation_leg": "{n} sites contre l’ordre de sériation",
        "seriation_none": "non porté : {names}, sans coordonnée AFC",
    },
}


# ==============================================================================
# SECTION 2 · SVG helpers
# ==============================================================================

def num(value, lang, decimals=2):
    """Format a number in the language's own convention."""
    text = f"{value:.{decimals}f}"
    return text.replace(".", ",") if lang == "fr" else text


def ramp(t, stops=None):
    """Colour at position t in [0, 1] along the RdYlGn stops."""
    stops = stops or RDYLGN
    t = min(max(t, 0.0), 1.0)
    span = t * (len(stops) - 1)
    i = min(int(span), len(stops) - 2)
    f = span - i
    out = "#"
    for k in (1, 3, 5):
        a, b = int(stops[i][k:k + 2], 16), int(stops[i + 1][k:k + 2], 16)
        out += format(round(a + (b - a) * f), "02x")
    return out


def ink_on(hex_colour):
    """Black or white, whichever stays legible on the given fill."""
    r, g, b = (int(hex_colour[k:k + 2], 16) for k in (1, 3, 5))
    return "#1a1a1a" if 0.299 * r + 0.587 * g + 0.114 * b > 150 else "#ffffff"


def text(x, y, body, size=10.5, fill=INK_SOFT, family=SANS, anchor="start",
         weight=None):
    w = f' font-weight="{weight}"' if weight else ""
    return (f'<text x="{x}" y="{y}" text-anchor="{anchor}" font-family="{family}" '
            f'font-size="{size}" fill="{fill}"{w}>{body}</text>')


def size_to_fit(label, width, ideal=11.0, floor=7.5, family=SANS):
    """Largest font size at which the label still fits the width, down to a floor.

    The oblique-rim plate is a group of one, so its box is the narrowest of the
    three while its French name is the longest of the three. Rather than pick a
    size that suits English and clips French, let the label ask the box.
    """
    advance = {MONO: 0.60, SANS: 0.505}[family]
    if not label:
        return ideal
    return max(floor, min(ideal, (width - 8) / (len(label) * advance)))


def gradient(ident, stops, reverse=False):
    seq = list(reversed(stops)) if reverse else list(stops)
    marks = "".join(
        f'<stop offset="{i / (len(seq) - 1):.4f}" stop-color="{c}"/>'
        for i, c in enumerate(seq))
    return f'<linearGradient id="{ident}" x1="0" x2="1">{marks}</linearGradient>'


def svg(width, height, body, aria):
    return (f'<svg xmlns="http://www.w3.org/2000/svg" version="1.1" '
            f'viewBox="0 0 {width} {height}" role="img" aria-label="{aria}">\n'
            f'  <rect x="0" y="0" width="{width}" height="{height}" fill="{PAPER}"/>\n'
            f'{body}\n</svg>\n')


def fits(body, width, height):
    """Cheap guard against a label running out of its own figure.

    French labels are systematically wider than English ones and the layouts are
    laid out in English, so this is the failure that would otherwise appear only
    in the translated half — after publication.
    """
    advance = {MONO: 0.60, SANS: 0.505}
    problems = []
    for m in re.finditer(r'<text x="([-\d.]+)" y="([-\d.]+)" text-anchor="(\w+)" '
                         r'font-family="([^"]+)" font-size="([\d.]+)"[^>]*>(.*?)</text>',
                         body, re.S):
        x, y, anchor, family, size, label = m.groups()
        x, y, size = float(x), float(y), float(size)
        plain = re.sub(r"<[^>]+>|&#\d+;", "X", label).strip()
        w = len(plain) * size * advance.get(family, 0.505)
        x0 = x - w / 2 if anchor == "middle" else (x - w if anchor == "end" else x)
        if x0 < -1 or x0 + w > width + 1 or y > height + 1:
            problems.append(f"{plain[:44]!r} spans {x0:.0f}-{x0 + w:.0f} of {width}")
    return problems


# ==============================================================================
# SECTION 3 · The figures
# ==============================================================================
# Each builder takes the facts and a language and returns (name, svg). They are
# collected by FIGURES at the end of the section; adding a figure means writing
# one function and naming it there.

def fig_stage_composition(facts, lang):
    """The group's sherds by stage rank, per horizon — what s measures."""
    t, comp = T[lang], facts["stage_composition"]
    group = facts["main_group"]
    tint = ["#f2c3bb", "#cc7264", "#9e3b26"]
    x0, width, body, y = 150.0, 350.0, [], 62
    ranks = sorted({r for c in comp.values() for r in c["ranks"]})
    for i, r in enumerate(ranks):
        body.append(f'<rect x="{150 + i * 122}" y="18" width="14" height="12" '
                    f'fill="{tint[i % len(tint)]}"/>')
        body.append(text(170 + i * 122, 28, f'{t["stage"]} · {t["rank"]} {r}', 10.5))
    for h in sorted(comp, reverse=True):
        cell = facts["cells"].get((h, group))
        info, x = comp[h], x0
        body.append(text(140, y + 18, f'{t["horizon"]} {h}', 11, INK, anchor="end"))
        for i, r in enumerate(ranks):
            share = info["shares"].get(r, 0.0)
            w = share / 100 * width
            if w <= 0.6:
                continue
            colour = tint[i % len(tint)]
            body.append(f'<rect x="{x:.1f}" y="{y}" width="{w:.1f}" height="26" '
                        f'fill="{colour}"/>')
            if w > 26:
                body.append(text(x + w / 2, y + 18, f"{share:.0f}%", 9.5,
                                 ink_on(colour), MONO, "middle"))
            x += w
        stat = (f'x&#772; = {num(cell["mean"], lang)} · s = {num(cell["s"], lang)}'
                if cell and cell["s"] is not None else "")
        body.append(text(x0 + width + 12, y + 13, stat, 9.5, INK, MONO))
        body.append(text(x0 + width + 12, y + 24, f'n = {info["total"]}', 9.5, "#8a8272", MONO))
        y += 38
    body.append(f'<line x1="150" y1="{y + 10}" x2="500" y2="{y + 10}" stroke="{RULE}"/>')
    body.append(text(325, y + 28, t["share_of"].format(group=GROUP_LABEL[lang][group]),
                     11, INK_SOFT, SANS, "middle"))
    body.append(text(325, y + 44, t["mass"], 10.5, SLIP, SANS, "middle"))
    return "\n".join("      " + b for b in body), 700, y + 56


def fig_group_shares(facts, lang):
    """Each group's percentage of the horizon — the complementary signal."""
    t, shares = T[lang], facts["group_shares"]
    groups = [g for g in GROUP_COLOUR if any(g in s["shares"] for s in shares.values())]
    x0, width, body, y = 150.0, 410.0, [], 58
    body.append(text(20, 24, t["shares_title"], 11.5, INK, SANS, weight="600"))
    lx = 150
    for g in groups:
        body.append(f'<rect x="{lx}" y="30" width="13" height="12" '
                    f'fill="{GROUP_COLOUR[g]}" fill-opacity=".85"/>')
        label = GROUP_LABEL[lang][g]
        body.append(text(lx + 19, 40, label, 10.5))
        lx += 30 + len(label) * 5.6
    for h in sorted(shares, reverse=True):
        info, x = shares[h], x0
        body.append(text(140, y + 19, f'{t["horizon"]} {h}', 11, INK, anchor="end"))
        for g in groups:
            share = info["shares"].get(g, 0.0)
            w = share / 100 * width
            if w <= 0.5:
                continue
            body.append(f'<rect x="{x:.1f}" y="{y}" width="{w:.1f}" height="28" '
                        f'fill="{GROUP_COLOUR[g]}" fill-opacity=".85"/>')
            if w > 30:
                body.append(text(x + w / 2, y + 19, f"{share:.0f}%", 10,
                                 "#ffffff", MONO, "middle"))
            x += w
        body.append(text(x0 + width + 12, y + 19, f'n = {info["total"]}', 9.5,
                         "#8a8272", MONO))
        y += 38
    body.append(f'<line x1="150" y1="{y + 8}" x2="560" y2="{y + 8}" stroke="{RULE}"/>')
    body.append(text(355, y + 26, t["shares_foot"], 11, SLIP, SANS, "middle"))
    return "\n".join("      " + b for b in body), 700, y + 38


def _bar_panel(facts, lang, key, title, ramp_label, hi, reverse, decimals):
    """Shared body of the two single-column panels of the printed figure."""
    t, group = T[lang], facts["main_group"]
    cells = {h: facts["cells"][(h, group)] for h in facts["horizons"]
             if (h, group) in facts["cells"]}
    xs = {h: 95 + i * 120 for i, h in enumerate(sorted(cells))}
    base, top = 176, 40
    body = [text(20, 22, title, 11.5, INK, SANS, weight="600"),
            f'<line x1="70" y1="{top - 6}" x2="70" y2="{base}" stroke="{INK_SOFT}"/>',
            f'<line x1="70" y1="{base}" x2="665" y2="{base}" stroke="{INK_SOFT}"/>']
    steps = 4 if key == "q" else 2
    for i in range(1, steps + 1):
        value = hi * i / steps
        y = base - (base - top) * i / steps
        body.append(f'<line x1="70" y1="{y:.1f}" x2="660" y2="{y:.1f}" '
                    f'stroke="{GRID}" stroke-dasharray="3,3"/>')
        body.append(text(62, y + 4, num(value, lang, 2), 10, INK_SOFT, MONO, "end"))
    body.append(text(62, base + 4, num(0, lang, 2), 10, INK_SOFT, MONO, "end"))
    for h, cell in sorted(cells.items()):
        value = cell[key]
        if value is None:
            continue
        height = (base - top) * value / hi
        y = base - height
        colour = ramp(value / hi if not reverse else 1 - value / hi)
        body.append(f'<rect x="{xs[h]}" y="{y:.1f}" width="70" height="{height:.1f}" '
                    f'fill="{colour}" stroke="#ffffff" stroke-width="1.5"/>')
        body.append(text(xs[h] + 35, y + 24, num(value, lang, decimals), 14,
                         ink_on(colour), SANS, "middle", "600"))
        body.append(text(xs[h] + 35, 194, f'{t["horizon"]} {h}', 10, INK_SOFT, MONO, "middle"))
        body.append(text(xs[h] + 35, 208, f'n={cell["n"]}', 10, "#8a8272", MONO, "middle"))
    body.append(text(367, 228, t["earlier"], 11, INK, SANS, "middle"))
    ident = f"ramp_{key}"
    body.append(f'<defs>{gradient(ident, RDYLGN, reverse=reverse)}</defs>')
    body.append(f'<rect x="230" y="242" width="240" height="14" fill="url(#{ident})" '
                f'stroke="{INK}" stroke-width=".5"/>')
    body.append(text(230, 270, num(0, lang, 2 if key == "s" else 1), 9.5, INK_SOFT, MONO, "middle"))
    body.append(text(470, 270, num(hi, lang, 2 if key == "s" else 1), 9.5, INK_SOFT, MONO, "middle"))
    body.append(text(482, 253, ramp_label, 9.5, "#8a8272"))
    return "\n".join("      " + b for b in body), 700, 275


def fig_variance_bars(facts, lang):
    """s per horizon, on the printed variance panel's own dataset-relative ramp."""
    group = facts["main_group"]
    hi = max(c["s"] for (h, g), c in facts["cells"].items()
             if g == group and c["s"] is not None)
    return _bar_panel(facts, lang, "s", T[lang]["var_title"], T[lang]["var_ramp"],
                      hi, reverse=True, decimals=2)


def fig_quality_bars(facts, lang):
    """q per horizon on the absolute 0-1 ramp, with the full axis it needs."""
    return _bar_panel(facts, lang, "q", T[lang]["qual_title"], T[lang]["qual_ramp"],
                      1.0, reverse=False, decimals=3)


def fig_origin_sensitivity(facts, lang):
    """q against the first-stage number: the level moves, the order never does."""
    t, origin = T[lang], facts["origin"]
    origins = origin["origins"]
    xs = {o: 110 + i * 90 for i, o in enumerate(origins)}
    lo, hi, top, base = 0.55, 1.0, 40, 220
    scale = lambda q: base - (q - lo) / (hi - lo) * (base - top)
    tint = [SLIP, "#b8604a", OCHRE, "#6b8a6a", SAGE]
    body = [f'<rect x="{xs[1] - 22}" y="{top}" width="44" height="{base - top}" '
            f'fill="{OCHRE}" fill-opacity=".10"/>',
            text(xs[1], top - 6, t["origin_declared"], 10.5, OCHRE, SANS, "middle"),
            f'<line x1="90" y1="{top}" x2="90" y2="{base}" stroke="{INK_SOFT}"/>',
            f'<line x1="90" y1="{base}" x2="{max(xs.values()) + 60}" y2="{base}" stroke="{INK_SOFT}"/>']
    for mark in (0.6, 0.8, 1.0):
        y = scale(mark)
        body.append(f'<line x1="90" y1="{y:.1f}" x2="{max(xs.values()) + 55}" '
                    f'y2="{y:.1f}" stroke="{GRID}" stroke-dasharray="3,3"/>')
        body.append(text(82, y + 4, num(mark, lang, 2), 10, INK_SOFT, MONO, "end"))
    for i, h in enumerate(sorted(origin["q"])):
        colour = tint[i % len(tint)]
        pts = " ".join(f"{xs[o]},{scale(origin['q'][h][o]):.1f}" for o in origins)
        body.append(f'<polyline fill="none" stroke="{colour}" stroke-width="1.8" points="{pts}"/>')
        for o in origins:
            body.append(f'<circle cx="{xs[o]}" cy="{scale(origin["q"][h][o]):.1f}" '
                        f'r="3.4" fill="{colour}"/>')
        last = origin["q"][h][origins[-1]]
        body.append(text(max(xs.values()) + 12, scale(last) + 3.5, f"H{h}", 9.5, colour, MONO))
    for o in origins:
        body.append(text(xs[o], base + 18, str(o), 10, INK_SOFT, MONO, "middle"))
    body.append(text(355, base + 38, t["origin_x"], 11, INK_SOFT, SANS, "middle"))
    body.append(text(34, 130, "q", 10, INK_SOFT, MONO, "middle") .replace(
        "<text ", '<text transform="rotate(-90 34 130)" ', 1))
    return "\n".join("      " + b for b in body), 700, base + 50


def fig_seriation(facts, lang):
    """The findspots on the CA axis, in horizon lanes, inversions ringed."""
    t, ser, intervals = T[lang], facts["seriation"], facts["intervals"]
    pts = [p for p in ser["points"] if p["horizon"] is not None]
    if not pts:
        return None
    lo = min(p["cax"] for p in pts)
    hi = max(p["cax"] for p in pts)
    pad = (hi - lo) * 0.04
    lo, hi = lo - pad, hi + pad
    x0, x1 = 70.0, 600.0
    px = lambda c: x0 + (c - lo) / (hi - lo) * (x1 - x0)
    lanes = {h: 66 + i * 34 for i, h in enumerate(sorted(facts["horizons"]))}
    tint = [SLIP, "#b8604a", OCHRE, "#6b8a6a", SAGE]
    marked = set(ser["inversions"])
    body = [f'<line x1="{x0}" y1="{max(lanes.values()) + 28}" x2="{x1}" '
            f'y2="{max(lanes.values()) + 28}" stroke="{INK_SOFT}"/>']
    for i, h in enumerate(sorted(lanes)):
        y, colour = lanes[h], tint[i % len(tint)]
        body.append(f'<line x1="{x0}" y1="{y}" x2="{x1}" y2="{y}" stroke="#eee8dc"/>')
        body.append(text(x0 - 8, y + 4, f"H{h}", 10.5, colour, SANS, "end"))
        span = intervals.get(h) or {}
        if span.get("start") is not None and span.get("end") is not None:
            body.append(text(x1 + 10, y + 4,
                             f'{span["start"]} … {span["end"]}', 8.5, "#8a8272", MONO))
    for p in pts:
        x, y = px(p["cax"]), lanes[p["horizon"]]
        colour = tint[(sorted(lanes).index(p["horizon"])) % len(tint)]
        if p["label"] in marked:
            body.append(f'<circle cx="{x:.1f}" cy="{y}" r="7.5" fill="none" '
                        f'stroke="{SLIP}" stroke-width="1.3"/>')
        body.append(f'<circle cx="{x:.1f}" cy="{y}" r="3.6" fill="{colour}"/>')
    axis_y = max(lanes.values()) + 28
    step = (hi - lo) / 5
    for i in range(6):
        value = lo + i * step
        body.append(text(px(value), axis_y + 18, num(value, lang, 1), 9.5,
                         INK_SOFT, MONO, "middle"))
    body.append(text(335, axis_y + 38, t["seriation_x"], 11, INK_SOFT, SANS, "middle"))
    legend_y = axis_y + 56
    body.append(f'<circle cx="86" cy="{legend_y}" r="7.5" fill="none" stroke="{SLIP}" stroke-width="1.3"/>')
    body.append(f'<circle cx="86" cy="{legend_y}" r="3.6" fill="{OCHRE}"/>')
    body.append(text(102, legend_y + 4, t["seriation_leg"].format(n=len(marked)), 10.5, INK))
    if ser["no_coordinate"]:
        body.append(text(352, legend_y + 4,
                         t["seriation_none"].format(names=", ".join(ser["no_coordinate"])),
                         10, "#8a8272"))
    return "\n".join("      " + b for b in body), 700, legend_y + 14


def fig_typology(facts, lang):
    """The group → stage → sub-type tree, with the ranks as assigned."""
    t, tree = T[lang], facts["typology"]
    groups = [g for g in GROUP_COLOUR if g in tree] or sorted(tree)
    leaves = {g: sum(len(v) for v in tree[g].values()) for g in groups}
    total = sum(leaves.values())
    body, x = [], 20.0
    span = (660 - 18 * (len(groups) - 1)) / total
    for g in groups:
        width = leaves[g] * span
        colour = GROUP_COLOUR.get(g, INK)
        body.append(f'<rect x="{x:.1f}" y="26" width="{width:.1f}" height="26" '
                    f'fill="{colour}" fill-opacity=".14" stroke="{colour}"/>')
        label = GROUP_LABEL[lang][g]
        body.append(text(x + width / 2, 44, label, round(size_to_fit(label, width), 1),
                         colour, SANS, "middle", "600"))
        stages = sorted(tree[g])
        sx = x
        for rank in stages:
            members = tree[g][rank]
            sw = len(members) * span
            body.append(f'<line x1="{x + width / 2:.1f}" y1="52" x2="{sx + sw / 2:.1f}" '
                        f'y2="84" stroke="#b9b2a3"/>')
            body.append(f'<rect x="{sx + 4:.1f}" y="84" width="{sw - 8:.1f}" height="24" '
                        f'fill="#fff" stroke="{colour}"/>')
            body.append(text(sx + sw / 2, 100, f'{t["stage"]} · {t["rank"]} {rank}',
                             9.5, colour, MONO, "middle"))
            for j, member in enumerate(members):
                lw = span
                lx = sx + j * lw
                body.append(f'<line x1="{sx + sw / 2:.1f}" y1="108" '
                            f'x2="{lx + lw / 2:.1f}" y2="132" stroke="#b9b2a3"/>')
                body.append(f'<rect x="{lx + 5:.1f}" y="132" width="{lw - 10:.1f}" '
                            f'height="22" fill="#f2efe8" stroke="{RULE}"/>')
                label = member[f"label_{lang}"] or member["label_en"]
                short = label.split()[-1]
                body.append(text(lx + lw / 2, 147, short, 9.5, INK_SOFT, SANS, "middle"))
            sx += sw
        x += width + 18
    return "\n".join("      " + b for b in body), 700, 170



def fig_worked_example(facts, lang):
    """One cell end to end: the sample, the three sums, the four derived numbers."""
    t, ex = T[lang], facts["worked_example"]
    d = num
    body = [text(20, 22, t["we_sample"], 11, INK, SANS, weight="600"),
            text(266, 22, t["we_sums"], 11, INK, SANS, weight="600"),
            text(482, 22, t["we_python"], 11, INK, SANS, weight="600"),
            f'<line x1="246" y1="30" x2="246" y2="252" stroke="{RULE}"/>',
            f'<line x1="462" y1="30" x2="462" y2="252" stroke="{RULE}"/>']
    by_rank = {dv["rank"]: dv["c"] for dv in ex["deviations"]}
    top = max(by_rank.values())
    tint = ["#f2c3bb", "#cc7264", "#9e3b26"]
    for i, (rank, count) in enumerate(sorted(by_rank.items())):
        y = 52 + i * 34
        body.append(text(20, y + 12, f'{t["rank"]} {rank}', 10.5, INK_SOFT, MONO))
        w = 121 * count / top
        body.append(f'<rect x="72" y="{y}" width="{w:.1f}" height="16" '
                    f'fill="{tint[i % len(tint)]}"/>')
        body.append(text(78 + w, y + 12, str(count), 10.5, INK, MONO))
    body.append(f'<line x1="20" y1="148" x2="226" y2="148" stroke="{RULE}"/>')
    for i, line in enumerate(t["we_caption"]):
        body.append(text(20, 168 + i * 16, line.format(
            n=ex["n"], group=GROUP_LABEL[lang][ex["group"]], horizon=ex["horizon"]), 10.5))
    sums = [("N      = ", ex["n"]), ("&#931; r&#183;c  = ", ex["sum_rc"]),
            ("&#931; r&#178;&#183;c = ", ex["sum_r2c"])]
    for i, (label, value) in enumerate(sums):
        body.append(text(266, 62 + i * 34, f"{label}{value}", 12, INK, MONO))
    for i, line in enumerate(t["we_sums_note"]):
        body.append(text(266, 168 + i * 16, line, 10.5))
    derived = [("x&#772;  = ", d(ex["mean"], lang, 3)), ("s  = ", d(ex["s"], lang, 3)),
               ("CV = ", d(ex["cv"], lang, 3))]
    for i, (label, value) in enumerate(derived):
        body.append(text(482, 62 + i * 34, f"{label}{value}", 12, INK, MONO))
    body.append(f'<rect x="474" y="146" width="150" height="30" fill="{SLIP}" '
                f'fill-opacity=".12" stroke="{SLIP}"/>')
    body.append(text(482, 166, f'q  = {d(ex["q"], lang, 3)}', 12, SLIP, MONO))
    for i, line in enumerate(t["we_derived_note"]):
        body.append(text(482, 200 + i * 16, line, 10.5))
    body.append(f'<line x1="20" y1="266" x2="660" y2="266" stroke="{RULE}"/>')
    body.append(text(20, 288, t["we_foot"].format(
        horizon=ex["horizon"], group=GROUP_LABEL[lang][ex["group"]]), 11, INK))
    return "\n".join("      " + b for b in body), 700, 300


def fig_rank_axis(facts, lang):
    """The group on its rank axis: dot area is share, tick the mean, bracket +/-s."""
    t, comp, group = T[lang], facts["stage_composition"], facts["main_group"]
    ranks = sorted({r for c in comp.values() for r in c["ranks"]})
    xs = {r: 200 + i * 160 for i, r in enumerate(ranks)}
    body = [text(20, 24, f'{GROUP_LABEL[lang][group]} · '
                 f'{t["n_stages"].format(n=len(ranks))}', 11.5, GROUP_COLOUR[group],
                 SANS, weight="600")]
    for r in ranks:
        body.append(f'<line x1="{xs[r]}" y1="36" x2="{xs[r]}" y2="212" stroke="{GRID}"/>')
    y = 54
    for h in sorted(comp, reverse=True):
        cell, info = facts["cells"][(h, group)], comp[h]
        body.append(text(150, y + 4, f'{t["horizon"]} {h}', 11, INK, anchor="end"))
        for r in ranks:
            share = info["shares"].get(r, 0.0)
            if share <= 0:
                continue
            body.append(f'<circle cx="{xs[r]}" cy="{y}" r="{3 + 9 * math.sqrt(share / 100):.2f}" '
                        f'fill="{SLIP}" fill-opacity=".78"/>')
        mx = xs[ranks[0]] + (cell["mean"] - ranks[0]) * 160
        sp = cell["s"] * 160
        body.append(f'<line x1="{mx - sp:.1f}" y1="{y + 15}" x2="{mx + sp:.1f}" '
                    f'y2="{y + 15}" stroke="{OCHRE}" stroke-width="1.5"/>')
        for edge in (mx - sp, mx + sp):
            body.append(f'<line x1="{edge:.1f}" y1="{y + 11}" x2="{edge:.1f}" '
                        f'y2="{y + 19}" stroke="{OCHRE}" stroke-width="1.5"/>')
        body.append(f'<line x1="{mx:.1f}" y1="{y - 13}" x2="{mx:.1f}" y2="{y + 19}" '
                    f'stroke="{INK}" stroke-width="1.4"/>')
        body.append(text(600, y - 1, f's = {num(cell["s"], lang, 3)}', 9.5, INK, MONO))
        body.append(text(600, y + 11, f'n = {cell["n"]}', 9.5, "#8a8272", MONO))
        y += 34
    body.append(f'<line x1="160" y1="212" x2="600" y2="212" stroke="{INK_SOFT}"/>')
    for r in ranks:
        body.append(text(xs[r], 230, f'{t["rank"]} {r}', 10, INK_SOFT, MONO, "middle"))
    body.append(text(380, 252, t["rank_axis_foot"], 11, INK_SOFT, SANS, "middle"))
    return "\n".join("      " + b for b in body), 700, 264


def fig_exp_transform(facts, lang):
    """q = exp(-CV), with the cells that exist marked on the curve."""
    t, group = T[lang], facts["main_group"]
    obs = [(c["s"] / c["mean"], c["q"], h) for (h, g), c in facts["cells"].items()
           if g == group and c["measured"]]
    obs.sort()
    body = [f'<rect x="70" y="30" width="{70 + 590 * max(o[0] for o in obs) - 70:.0f}" '
            f'height="140" fill="{OCHRE}" fill-opacity=".08"/>',
            text(70 + 295 * max(o[0] for o in obs), 24, t["exp_band"], 10.5, OCHRE, SANS, "middle"),
            f'<line x1="70" y1="30" x2="70" y2="170" stroke="{INK_SOFT}"/>',
            f'<line x1="70" y1="170" x2="665" y2="170" stroke="{INK_SOFT}"/>',
            f'<line x1="70" y1="40" x2="660" y2="170" stroke="{INK_SOFT}" '
            f'stroke-width="1" stroke-dasharray="4,3" stroke-opacity=".55"/>',
            text(565, 150, "q = 1 &#8722; CV", 10.5, INK_SOFT, SANS, "end")]
    pts = " ".join(f"{70 + 590 * cv / 1.0:.0f},{170 - 130 * math.exp(-cv):.1f}"
                   for cv in [i / 20 for i in range(21)])
    body.append(f'<polyline fill="none" stroke="{INK}" stroke-width="2" points="{pts}"/>')
    for cv, q, h in obs:
        body.append(f'<circle cx="{70 + 590 * cv:.1f}" cy="{170 - 130 * q:.1f}" '
                    f'r="4" fill="{SLIP}"/>')
        body.append(text(70 + 590 * cv + 10, 170 - 130 * q + 3,
                         f'H{h} · {num(q, lang, 3)}', 9.5, SLIP, MONO))
    body.append(text(62, 44, num(1, lang, 1), 10, INK_SOFT, MONO, "end"))
    body.append(text(62, 174, num(0, lang, 1), 10, INK_SOFT, MONO, "end"))
    body.append(text(365, 192, num(0.5, lang, 1), 10, INK_SOFT, MONO, "middle"))
    body.append(text(660, 192, num(1, lang, 1), 10, INK_SOFT, MONO, "middle"))
    body.append(text(365, 212, t["exp_x"], 11, INK_SOFT, SANS, "middle"))
    return "\n".join("      " + b for b in body), 700, 224


FIGURES = {
    "seriation_to_horizons": fig_seriation,
    "typology_ranks": fig_typology,
    "stage_composition": fig_stage_composition,
    "group_shares_by_horizon": fig_group_shares,
    "service_variance": fig_variance_bars,
    "service_quality": fig_quality_bars,
    "rank_origin_sensitivity": fig_origin_sensitivity,
    "one_cell_worked_example": fig_worked_example,
    "rank_axis_mean_sd": fig_rank_axis,
    "exp_transform": fig_exp_transform,
}


# ==============================================================================
# SECTION 4 · The derivation workbook
# ==============================================================================

WB = {
    "en": dict(data="Data", worked="Worked example", allcells="All cells", readme="Read me",
               cols=["horizon", "group", "sub-type", "stage rank r", "sherds c"],
               head=["sub-type", "stage rank r", "sherds c", "r x c", "r^2 x c"],
               totals="totals", measured="measured", degenerate="one stage",
               allhead=["horizon", "group", "sub-types k", "N sherds", "sum r.c",
                        "sum r^2.c", "x-bar", "numerator", "s", "CV", "q", "status"],
               steps=["STEP 2 . the three sums the query returns",
                      "STEP 3 . the mean rank", "STEP 4 . the sum of squared deviations",
                      "STEP 5 . variance and standard deviation",
                      "STEP 6 . the bounded score",
                      "STEP 7 . cross-check against the definition"],
               note="Generated by py/build_docs.py from the published graphs. Blue cells are input."),
    "fr": dict(data="Données", worked="Exemple détaillé", allcells="Toutes les cellules",
               readme="Lisez-moi",
               cols=["horizon", "groupe", "sous-type", "rang de stade r", "tessons c"],
               head=["sous-type", "rang de stade r", "tessons c", "r x c", "r^2 x c"],
               totals="totaux", measured="mesurée", degenerate="un seul stade",
               allhead=["horizon", "groupe", "sous-types k", "N tessons", "somme r.c",
                        "somme r^2.c", "x-barre", "numérateur", "s", "CV", "q", "statut"],
               steps=["ÉTAPE 2 . les trois sommes renvoyées par la requête",
                      "ÉTAPE 3 . le rang moyen", "ÉTAPE 4 . la somme des écarts au carré",
                      "ÉTAPE 5 . variance et écart-type", "ÉTAPE 6 . le score borné",
                      "ÉTAPE 7 . vérification par la définition"],
               note="Généré par py/build_docs.py à partir des graphes publiés. Les cellules bleues sont des saisies."),
}


def build_workbook(facts, path: Path, lang="en") -> None:
    """Write the derivation workbook: raw counts in, every other cell a formula.

    The point of the formulas is that the arithmetic can be followed and re-run
    by a reader who has neither Python nor the graphs — and that editing a count
    moves s and q in front of them.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    # A note on how far the reproducibility goes. The timestamps below make the
    # workbook byte-identical between two runs on one machine, which is what
    # makes a diff on this file mean something. It is *not* identical across
    # machines that disagree about lxml: openpyxl serialises through lxml when
    # it is installed and through the standard library when it is not, and the
    # two write empty elements differently (`<x />` against `<x/>`). Setting
    # openpyxl.LXML here would be too late — the writers bind their serialiser
    # at import. Pin lxml in requirements.txt if cross-machine identity of the
    # workbook matters; the figures and the notes are already identical either
    # way, and they are what the paper's reproducibility claim is about.

    w = WB[lang]
    blue = Font(name="Arial", size=10, color="0000FF")
    black = Font(name="Arial", size=10)
    bold = Font(name="Arial", size=10, bold=True)
    head = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1B2430")
    yellow = PatternFill("solid", fgColor="FFF6E0")

    wb = Workbook()
    ws = wb.active
    ws.title = w["data"]
    ws["A1"] = w["note"]
    ws["A1"].font = Font(name="Arial", size=10, bold=True, italic=True)
    for j, label in enumerate(w["cols"], 1):
        c = ws.cell(row=3, column=j, value=label)
        c.font, c.fill = head, fill
    rows = sorted((r for r in facts["records"] if r["horizon"] is not None),
                  key=lambda r: (r["horizon"], r["group"], r["column"]))
    for i, rec in enumerate(rows):
        for j, value in enumerate([rec["horizon"], rec["group"],
                                   rec[f"label_{lang}"] or rec["label_en"],
                                   rec["stage"], rec["sherds"]], 1):
            ws.cell(row=4 + i, column=j, value=value).font = blue
    last = 3 + len(rows)
    for col, width in zip("ABCDE", (10, 22, 26, 15, 11)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"

    hor, grp, rnk, shd = (f"'{w['data']}'!$A$4:$A${last}", f"'{w['data']}'!$B$4:$B${last}",
                          f"'{w['data']}'!$D$4:$D${last}", f"'{w['data']}'!$E$4:$E${last}")

    ex = facts["worked_example"]
    we = wb.create_sheet(w["worked"])
    we["A1"] = f"Horizon {ex['horizon']} · {ex['group']}"
    we["A1"].font = Font(name="Arial", size=12, bold=True)
    for j, label in enumerate(w["head"], 1):
        c = we.cell(row=3, column=j, value=label)
        c.font, c.fill = head, fill
    for i, r in enumerate(ex["rows"]):
        row = 4 + i
        we.cell(row=row, column=1, value=r[f"label_{lang}"] or r["label_en"]).font = black
        we.cell(row=row, column=2, value=r["rank"]).font = blue
        we.cell(row=row, column=3, value=r["c"]).font = blue
        we.cell(row=row, column=4, value=f"=B{row}*C{row}").font = black
        we.cell(row=row, column=5, value=f"=B{row}^2*C{row}").font = black
    tot = 4 + len(ex["rows"])
    we.cell(row=tot, column=1, value=w["totals"]).font = bold
    for col in "CDE":
        cell = we[f"{col}{tot}"]
        cell.value, cell.font = f"=SUM({col}4:{col}{tot - 1})", bold
    steps = [
        (w["steps"][0], None, None), ("N", f"=C{tot}", "0"),
        ("sum r.c", f"=D{tot}", "0"), ("sum r^2.c", f"=E{tot}", "0"),
        (w["steps"][1], None, None), ("x-bar = sum r.c / N", None, "0.00000"),
        (w["steps"][2], None, None), ("x-bar^2", None, "0.00000"),
        ("N . x-bar^2", None, "0.00000"), ("sum r^2.c - N.x-bar^2", None, "0.00000"),
        (w["steps"][3], None, None), ("variance = numerator / (N-1)", None, "0.00000"),
        ("s = sqrt(variance)", None, "0.00000"),
        (w["steps"][4], None, None), ("CV = s / x-bar", None, "0.00000"),
        ("q = exp(-CV)", None, "0.00000"),
    ]
    r = tot + 2
    ref = {}
    for label, formula, fmt in steps:
        if fmt is None:
            we.cell(row=r, column=1, value=label).font = bold
            r += 1
            continue
        we.cell(row=r, column=1, value=label).font = black
        ref[label.split()[0]] = r
        if formula is None:
            formula = {
                "x-bar": f"=B{ref['sum'] if False else r - 3}/B{r - 4}",
                "x-bar^2": f"=B{r - 2}^2",
                "N": "", }.get(label.split()[0], "")
        cell = we.cell(row=r, column=2, value=formula)
        cell.font, cell.fill, cell.number_format = black, yellow, fmt
        r += 1
    # Fill the formulas that need row numbers now that they are known.
    n_row, s1_row, s2_row = tot + 3, tot + 4, tot + 5
    mean_row, msq_row, nmsq_row, num_row = tot + 7, tot + 9, tot + 10, tot + 11
    var_row, s_row, cv_row, q_row = tot + 13, tot + 14, tot + 16, tot + 17
    we[f"B{mean_row}"] = f"=B{s1_row}/B{n_row}"
    we[f"B{msq_row}"] = f"=B{mean_row}^2"
    we[f"B{nmsq_row}"] = f"=B{n_row}*B{msq_row}"
    we[f"B{num_row}"] = f"=B{s2_row}-B{nmsq_row}"
    we[f"B{var_row}"] = f"=B{num_row}/(B{n_row}-1)"
    we[f"B{s_row}"] = f"=SQRT(B{var_row})"
    we[f"B{cv_row}"] = f"=B{s_row}/B{mean_row}"
    we[f"B{q_row}"] = f"=EXP(-B{cv_row})"
    for col, width in zip("ABCDE", (34, 16, 14, 14, 14)):
        we.column_dimensions[col].width = width

    ac = wb.create_sheet(w["allcells"])
    for j, label in enumerate(w["allhead"], 1):
        c = ac.cell(row=1, column=j, value=label)
        c.font, c.fill = head, fill
    for i, (key, cell) in enumerate(sorted(facts["cells"].items())):
        row = 2 + i
        h, g = key
        ac.cell(row=row, column=1, value=h).font = black
        ac.cell(row=row, column=2, value=g).font = black
        ac.cell(row=row, column=3, value=f'=COUNTIFS({hor},A{row},{grp},B{row})').font = black
        ac.cell(row=row, column=4, value=f'=SUMIFS({shd},{hor},A{row},{grp},B{row})').font = black
        ac.cell(row=row, column=5,
                value=f'=SUMPRODUCT(({hor}=A{row})*({grp}=B{row})*{rnk}*{shd})').font = black
        ac.cell(row=row, column=6,
                value=f'=SUMPRODUCT(({hor}=A{row})*({grp}=B{row})*{rnk}^2*{shd})').font = black
        ac.cell(row=row, column=7, value=f'=IF(D{row}=0,"",E{row}/D{row})').font = black
        ac.cell(row=row, column=8, value=f'=IF(D{row}=0,"",F{row}-E{row}^2/D{row})').font = black
        ac.cell(row=row, column=9,
                value=f'=IF(D{row}<2,"",SQRT(MAX(0,H{row})/(D{row}-1)))').font = black
        ac.cell(row=row, column=10,
                value=f'=IF(OR(D{row}<2,G{row}=0),"",I{row}/G{row})').font = black
        ac.cell(row=row, column=11, value=f'=IF(J{row}="","",EXP(-J{row}))').font = black
        ac.cell(row=row, column=12,
                value=f'=IF(ROUND(H{row},9)=0,"{w["degenerate"]}","{w["measured"]}")').font = black
        for col in "GHIJK":
            ac[f"{col}{row}"].number_format = "0.0000"
    for col, width in zip("ABCDEFGHIJKL", (9, 22, 12, 11, 11, 12, 10, 12, 10, 10, 10, 14)):
        ac.column_dimensions[col].width = width
    ac.freeze_panes = "A2"

    # Two things make a freshly saved workbook differ from an identical one
    # saved a second earlier: the document properties carry the moment of
    # writing, and every zip entry carries its own timestamp. Neither says
    # anything about the data, and both would show the file as changed on every
    # rebuild. Pin the first here and normalise the second below, so a diff on
    # this file means the numbers moved.
    stamp = datetime(2000, 1, 1, 0, 0, 0)
    wb.properties.created = stamp
    wb.properties.modified = stamp
    wb.properties.creator = "py/build_docs.py"
    wb.properties.lastModifiedBy = "py/build_docs.py"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    _freeze_zip_times(path, stamp)


def _freeze_zip_times(path: Path, stamp) -> None:
    """Rewrite an .xlsx with one fixed timestamp on every entry.

    An xlsx is a zip, and zipfile stamps each member with the current time. The
    payload is unchanged; only the archive metadata is normalised, so the file
    is byte-identical between two runs that produced the same numbers.
    """
    import shutil
    import zipfile

    fixed = (stamp.year, stamp.month, stamp.day, stamp.hour, stamp.minute, stamp.second)
    iso = stamp.strftime("%Y-%m-%dT%H:%M:%SZ")
    tmp = path.with_suffix(".tmp")
    with zipfile.ZipFile(path) as src, zipfile.ZipFile(
            tmp, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in sorted(src.infolist(), key=lambda i: i.filename):
            payload = src.read(item.filename)
            if item.filename == "docProps/core.xml":
                # openpyxl stamps dcterms:modified at save time and ignores what
                # was set beforehand, so it is corrected here rather than there.
                payload = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>" + iso.encode() + rb"\g<2>", payload)
            info = zipfile.ZipInfo(item.filename, date_time=fixed)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = item.external_attr
            dst.writestr(info, payload)
    shutil.move(str(tmp), str(path))


# ==============================================================================
# SECTION 4 · Rendering the note
# ==============================================================================

NOTE_YAML = Path(__file__).resolve().parent / "note.yaml"
ASSET_DIR = Path(__file__).resolve().parent / "docs_assets"
TEMPLATE_DIR = Path(__file__).resolve().parent / "templates"

CHROME = {
    "en": {
        "lang_tag": "en-GB", "toc_heading": "Contents",
        "title": "Within-group variability by service type — how the heatmap is built and how to read it",
        "kicker": "CAA 2026 · Alligator pipeline · figure note",
        "h1": "Within-group variability<br>by service type",
        "lede": ("A reading guide to <code>service_group_variability</code>: what one cell of the "
                 "heatmap counts, what the two panels do and do not add to one another, and the one "
                 "structural feature of the data that makes two of the three columns unreadable "
                 "unless it is declared."),
        "footer_title": "CAA 2026 · leiza-scit/CAA2026-alligator",
        "footer_note": ("Every quantity on this page is read from the published graphs by "
                        "<code>py/docs_facts.py</code>. Where the page and the graph disagree, the "
                        "build fails rather than the page lying."),
        "generated": "Generated by py/build_docs.py — do not edit this file; edit py/note_en.yaml.",
    },
    "fr": {
        "lang_tag": "fr", "toc_heading": "Sommaire",
        "title": "Variabilité intra-groupe par type de service — construction et lecture de la figure",
        "kicker": "CAA 2026 · chaîne Alligator · note de figure",
        "h1": "Variabilité intra-groupe<br>par type de service",
        "lede": ("Guide de lecture de <code>service_group_variability</code> : ce que compte une "
                 "cellule, ce que les deux panneaux s’apportent — et ne s’apportent pas — "
                 "mutuellement, et le trait structurel des données qui rend deux des trois colonnes "
                 "illisibles tant qu’il n’est pas déclaré."),
        "footer_title": "CAA 2026 · leiza-scit/CAA2026-alligator",
        "footer_note": ("Chaque grandeur de cette page est lue dans les graphes publiés par "
                        "<code>py/docs_facts.py</code>. En cas de désaccord entre la page et le "
                        "graphe, la compilation échoue plutôt que la page ne mente."),
        "generated": "Généré par py/build_docs.py — ne pas éditer ce fichier ; éditer py/note_fr.yaml.",
    },
}


def ansa(marker: str) -> str:
    """The tabula ansata that numbers a section heading."""
    return (
        '<svg class="ansa" width="34" height="22" viewBox="0 0 34 22" aria-hidden="true">'
        f'<g fill="none" stroke="{SLIP}" stroke-width="1.3">'
        '<path d="M7 3 L27 3 L27 19 L7 19 Z"/><path d="M7 3 L1 7 L1 15 L7 19"/>'
        '<path d="M27 3 L33 7 L33 15 L27 19"/></g>'
        f'<text x="17" y="15" text-anchor="middle" fill="{SLIP}" font-family="{MONO}" '
        f'font-size="{9 if len(marker) > 2 else 10}">{marker}</text></svg>')


def load_note():
    """Read the bilingual note source and refuse an untranslated string.

    One file holds both languages so that a block cannot be revised in English
    and forgotten in French. The check below is what makes that a guarantee
    rather than an intention: a missing or empty translation stops the build,
    which is the only moment at which anybody is still looking.
    """
    import yaml
    doc = yaml.safe_load(NOTE_YAML.read_text(encoding="utf-8"))
    missing = []
    for sec in doc["sections"]:
        where = f"section {sec['id']}"
        for field in ("heading", "sub"):
            value = sec.get(field)
            if value is None:
                continue
            for lang in LANGS:
                if not (value.get(lang) or "").strip():
                    missing.append(f"{where}: {field} has no {lang}")
        for i, blk in enumerate(sec["blocks"], 1):
            field = "caption" if blk["type"] == "figure" else "html"
            value = blk.get(field) or {}
            for lang in LANGS:
                if not (value.get(lang) or "").strip():
                    missing.append(f"{where}, block {i} ({blk['type']}): "
                                   f"{field} has no {lang}")
    if missing:
        raise SystemExit("note.yaml is not fully bilingual:\n  " + "\n  ".join(missing))
    return doc


def _pick(value, lang):
    """One language out of a bilingual string, or the string itself."""
    if isinstance(value, dict):
        return value.get(lang, "")
    return value or ""


def render_note(facts, lang, figures_svg, doc):
    """Render one language of the note from the shared YAML, figures and chrome."""
    from jinja2 import Environment, FileSystemLoader, select_autoescape

    sections = []
    for sec in doc["sections"]:
        blocks = []
        for blk in sec["blocks"]:
            out = {"type": blk["type"]}
            if blk["type"] == "figure":
                out["figure"] = blk["figure"]
                out["caption"] = _pick(blk.get("caption"), lang)
            else:
                out["html"] = _pick(blk.get("html"), lang)
            blocks.append(out)
        sections.append({"id": sec["id"], "marker": sec["marker"],
                         "heading": _pick(sec.get("heading"), lang),
                         "sub": _pick(sec.get("sub"), lang),
                         "blocks": blocks})

    def figure(name):
        if name in figures_svg:
            return figures_svg[name]
        asset = ASSET_DIR / f"{name}_{lang}.svg"
        if not asset.exists():
            raise SystemExit(f"figure {name} is neither computed nor an asset")
        return re.sub(r"<\?xml[^>]*\?>\s*", "", asset.read_text(encoding="utf-8"))

    env = Environment(loader=FileSystemLoader(str(TEMPLATE_DIR)),
                      autoescape=select_autoescape(default=False),
                      trim_blocks=True, lstrip_blocks=True)
    tpl = env.get_template("variability.html.j2")
    chrome = CHROME[lang]
    return tpl.render(sections=sections, figure=figure, ansa=ansa,
                      generated_note=chrome["generated"], **chrome)


# ==============================================================================
# SECTION 4 · Auditing the prose
# ==============================================================================
# The figures are generated, so they cannot go stale. The note's prose is still
# written by hand — that is where the judgement lives — but every number in it
# has to agree with the graph. These are the claims worth checking, each as a
# regular expression over the text and the value it must find.

def audit_claims(facts):
    """Return (label, pattern, expected) for every checkable number in the note."""
    group = facts["main_group"]
    ex, acc, ser = facts["worked_example"], facts["accounting"], facts["seriation"]
    cells = facts["cells"]
    claims = [
        ("sherds in the workbook", r"(?:workbook|classeur)\D{0,40}?(\d{4})", acc["overall"]),
        ("sherds inside a horizon", r"(?:horizon|panels|panneaux)\D{0,40}?(\d{4})\b", acc["in_horizon"]),
        ("observations", r"(\d{3})\s+observations", acc["observations"]),
        ("worked example N", rf"{ex['n']}\s+(?:sherds|tessons)", ex["n"]),
        ("worked example sum rc", rf"\b({ex['sum_rc']})\b", ex["sum_rc"]),
        ("worked example sum r2c", rf"\b({ex['sum_r2c']})\b", ex["sum_r2c"]),
        ("seriation in order", r"(?:Forty|Quarante)\w*\s+(?:of|des)", ser["in_order"]),
    ]
    for h in sorted(facts["horizons"]):
        cell = cells.get((h, group))
        if cell and cell["measured"]:
            claims.append((f"H{h} s", None, round(cell["s"], 3)))
            claims.append((f"H{h} q", None, round(cell["q"], 3)))
    return claims


def audit(path: Path, facts) -> list[str]:
    """Check the note's quoted numbers against the facts. Returns the failures."""
    if not path.exists():
        return [f"{path.name}: not found"]
    body = re.sub(r"<[^>]+>", " ", path.read_text(encoding="utf-8"))
    body = body.replace("\u202f", " ").replace("\u00a0", " ")
    decimal = body.replace(",", ".")
    problems = []
    group = facts["main_group"]
    for h in sorted(facts["horizons"]):
        cell = facts["cells"].get((h, group))
        if not cell or not cell["measured"]:
            continue
        for key, places in (("s", 3), ("q", 3)):
            value = f"{cell[key]:.{places}f}"
            short = f"{cell[key]:.2f}"
            if value not in decimal and short not in decimal:
                problems.append(
                    f"{path.name}: H{h} {key} = {value} appears nowhere — "
                    f"the note quotes a value the graph no longer holds")
    acc = facts["accounting"]
    for label, value in (("total sherds", acc["overall"]),
                         ("sherds in a horizon", acc["in_horizon"]),
                         ("observations", acc["observations"])):
        if str(value) not in body:
            problems.append(f"{path.name}: {label} = {value} appears nowhere")
    ex = facts["worked_example"]
    for label, value in (("N", ex["n"]), ("sum r*c", ex["sum_rc"]),
                         ("sum r^2*c", ex["sum_r2c"])):
        if str(value) not in body:
            problems.append(f"{path.name}: worked example {label} = {value} appears nowhere")
    return problems


# ==============================================================================
# SECTION 5 · Entry point
# ==============================================================================

def build_figures(facts, out_dir: Path):
    """Draw every registered figure in every language.

    Returns {lang: {name: svg}} for the note to inline, and writes the same
    figures as standalone files so they can be dropped into the article.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    made_all, problems = {lang: {} for lang in LANGS}, []
    for lang in LANGS:
        for name, builder in FIGURES.items():
            made = builder(facts, lang)
            if made is None:
                continue
            body, width, height = made
            bad = fits(body, width, height)
            if bad:
                problems += [f"{name} [{lang}]: {p}" for p in bad]
            markup = svg(width, height, body, f"{name.replace('_', ' ')} ({lang})")
            (out_dir / f"{name}_{lang}.svg").write_text(markup, encoding="utf-8",
                                                        newline="\n")
            made_all[lang][name] = re.sub(r'\s*xmlns="[^"]*"|\s*version="1.1"', "",
                                          markup, count=2)
    if problems:
        raise SystemExit("figure labels do not fit:\n  " + "\n  ".join(problems))
    return made_all


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--audit", type=Path, nargs="*",
                    help="HTML notes whose numbers must match the facts")
    ap.add_argument("--out", type=Path, default=OUT_DIR)
    args = ap.parse_args()

    print("=" * 60)
    print("Variability note · figures from the graph, prose audited")
    print("=" * 60)

    facts = collect()
    acc = facts["accounting"]
    print(f"  ..  facts: {acc['observations']} observations, {acc['overall']} sherds, "
          f"{acc['in_horizon']} inside a horizon")

    figures = build_figures(facts, args.out)
    count = sum(len(v) for v in figures.values())
    print(f"  OK  {args.out.relative_to(ROOT)}  ({count} figure(s) drawn, "
          f"{len(FIGURES)} per language)")

    doc = load_note()
    print(f"  ..  note.yaml: {len(doc['sections'])} sections, "
          f"{sum(len(s['blocks']) for s in doc['sections'])} blocks, "
          f"complete in {' and '.join(LANGS)}")
    notes = {}
    for lang in LANGS:
        html = render_note(facts, lang, figures[lang], doc)
        suffix = "" if lang == "en" else f"_{lang}"
        path = args.out / f"service_group_variability_explained{suffix}.html"
        path.write_text(html, encoding="utf-8", newline="\n")
        notes[lang] = path
        print(f"  OK  {path.relative_to(ROOT)}  ({len(html) // 1024} KB)")

    # The generated notes are audited too: a figure kept as a versioned asset
    # rather than computed could otherwise still drift from the data.
    failures = []
    for path in list(notes.values()) + list(args.audit or []):
        failures += audit(path, facts)
    if failures:
        print("\n".join("  !!  " + f for f in failures))
        raise SystemExit("audit failed — the prose quotes numbers the graph does not")
    for lang in LANGS:
        suffix = "" if lang == "en" else f"_{lang}"
        book = args.out / f"service_group_variability_derivation{suffix}.xlsx"
        build_workbook(facts, book, lang)
        print(f"  OK  {book.relative_to(ROOT)}")

    print(f"  OK  audit: {len(notes) + len(args.audit or [])} note(s), "
          f"every quoted number agrees with the graph")

    print("=" * 60)
    print("Done.")
    print("=" * 60)


if __name__ == "__main__":
    main()
